# data.py
import matplotlib
matplotlib.use('Agg')  # 设置非交互式后端
# 添加中文字体支持
matplotlib.rcParams['font.sans-serif'] = ['SimHei']  # 指定默认字体为黑体
matplotlib.rcParams['axes.unicode_minus'] = False    # 正常显示负号
import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import seaborn as sns
import random
from datetime import datetime
import subprocess
import sys
import os
from pathlib import Path
import time  # 导入 time 模块
from concurrent.futures import ThreadPoolExecutor, as_completed  # 引入并发模块
import threading  # 添加 threading 模块导入
import logging  # 导入 logging 模块

# 设置日志
logging.basicConfig(
    level=logging.DEBUG,  # 设置为 DEBUG 级别以获取详细信息
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),  # 输出到终端
        logging.FileHandler("data.log", encoding='utf-8')  # 同时输出到日志文件
    ]
)

# 设置随机种子
random.seed(42)

# 定义项目根目录
PROJECT_ROOT = Path(__file__).parent.resolve()

# 定义基础输出目录
OUTPUTS_DIR = PROJECT_ROOT / "outputs"

# 创建基础输出目录（如果未创建）
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

# Configuration
USER_AGENT = os.getenv("USER_AGENT", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36")
COOKIE = os.getenv("COOKIE", "HMF_CI=eb51f625c54193e2c3da483f1f8a20849df1836e651910d5c13682ac94490c626fd7ee9b8958ab34c8abc8a8f7fc5dd52f24fdae80b7aaa45255ee8f42c1015380; 21_vq=3")
header = {
    "User-Agent": USER_AGENT,
    "cookie": COOKIE
}
URL = "https://www.cwl.gov.cn/cwl_admin/front/cwlkj/search/kjxx/findDrawNotice?"

# 获取当前时间戳，用于文件命名
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

# 输入验证函数
def get_valid_input(prompt, valid_options):
    while True:
        choice = input(prompt)
        if choice in valid_options:
            return choice
        else:
            print("无效的选择，请重新输入。")

# 获取用户输入的查询参数
query_type = get_valid_input(
    "1: 按期数查询\n2: 按期号查询\n3: 按日期查询\n请选择查询方式:",
    ['1', '2', '3']
)

params = {
    "name": "ssq",
    "issueCount": "",
    "issueStart": "",
    "issueEnd": "",
    "dayStart": "",
    "dayEnd": "",
    "pageNo": "1",
    "pageSize": "30",
    "week": "",
    "systemType": "PC"
}

if query_type == '1':
    params["issueCount"] = input("最近100期即100\n请输入期数：")
elif query_type == '2':
    params["issueStart"] = input("请输入起始期号：")
    params["issueEnd"] = input("请输入结束期号：")
elif query_type == '3':
    params["dayStart"] = input("请输入起始日期（格式：YYYY-MM-DD）：")
    params["dayEnd"] = input("请输入结束日期（格式：YYYY-MM-DD）：")

def fetch_data(url, params):
    try:
        response = requests.get(url, headers=header, params=params)
        response.raise_for_status()
        response.encoding = 'utf-8'
        return response.json()
    except requests.RequestException as e:
        logging.error(f"请求发生错误: {e}")
        return None
    except json.JSONDecodeError:
        logging.error("无法解码 JSON 数据")
        return None

# 并行获取数据的函数
def fetch_page(page_num):
    local_params = params.copy()
    local_params["pageNo"] = str(page_num)
    return fetch_data(URL, local_params)

# 开始计时
start_time = time.time()

# 获取第一页数据以确定总页数
res = fetch_data(URL, params)
if res is None:
    sys.exit()

pagenum = res.get("pageNum", 0)
logging.info(f"总页数: {pagenum}")

data = []
data_lock = threading.Lock()  # 锁用于线程安全地写入数据

def process_page(jsondata):
    page_data = []
    if jsondata is None or "result" not in jsondata:
        return page_data

    for result in jsondata["result"]:
        try:
            code = int(result["code"])
            red = [int(num) for num in result["red"].split(',')]
            blue = int(result["blue"])
            date = result["date"]
            week = result["week"]
            logging.debug(f"期数：{code} 红球：{red} 蓝球：{blue} 日期：{date} 星期：{week}")
            page_data.append([code, date, week] + red + [blue])
        except (ValueError, KeyError) as e:
            logging.error(f"数据解析错误: {e}")
            continue
    return page_data

# 使用线程池并行获取剩余页面的数据
with ThreadPoolExecutor(max_workers=10) as executor:
    # 提交所有页面的请求任务
    future_to_page = {executor.submit(fetch_page, i): i for i in range(1, int(pagenum) + 1)}
    for future in as_completed(future_to_page):
        page_num = future_to_page[future]
        try:
            jsondata = future.result()
            page_data = process_page(jsondata)
            with data_lock:
                data.extend(page_data)
        except Exception as e:
            logging.error(f"页面 {page_num} 处理时发生错误: {e}")

# 创建 DataFrame
columns = ["期数", "日期", "星期", "红球号码1", "红球号码2", "红球号码3", "红球号码4", "红球号码5", "红球号码6", "蓝球号码"]
df = pd.DataFrame(data, columns=columns)

# 检查和移除重复的期数
initial_count = len(df)
df.drop_duplicates(subset="期数", inplace=True)
final_count = len(df)
if initial_count != final_count:
    logging.warning(f"移除了 {initial_count - final_count} 个重复的期数。")

# 按期数排序
df.sort_values(by="期数", ascending=True, inplace=True)
df.reset_index(drop=True, inplace=True)
logging.info("数据已按期数升序排序。")

# 获取最新期数
if df.empty:
    logging.error("没有获取到任何数据。")
    sys.exit()

latest_period = df["期数"].max()
next_period = latest_period + 1
logging.info(f"最新期数: {latest_period}, 下一期数: {next_period}")

# 确定新的文件夹名称
def get_new_folder_name(base_dir, next_period):
    folder_name = f"{next_period}"
    folder_path = base_dir / folder_name
    if not folder_path.exists():
        return folder_path
    else:
        suffix = 1
        while True:
            new_folder_name = f"{next_period}_{suffix}"
            new_folder_path = base_dir / new_folder_name
            if not new_folder_path.exists():
                return new_folder_path
            suffix += 1

new_folder_path = get_new_folder_name(OUTPUTS_DIR, next_period)
logging.info(f"将文件保存到: {new_folder_path}")

# 定义输出目录
EXCELS_DIR = new_folder_path / "excels"
CHARTS_DIR = new_folder_path / "charts"
RED_PLOTS_DIR = CHARTS_DIR / "red_plots"
BLUE_PLOTS_DIR = CHARTS_DIR / "blue_plots"
PREDICTIONS_DIR = new_folder_path / "predictions"
DETAILS_DIR = PREDICTIONS_DIR / "details"
RESULTS_DIR = PREDICTIONS_DIR / "results"
HIGHLIGHTED_EXCELS_DIR = new_folder_path / "highlighted_excels"

# 创建所有必要的目录
for directory in [EXCELS_DIR, RED_PLOTS_DIR, BLUE_PLOTS_DIR, DETAILS_DIR, RESULTS_DIR, HIGHLIGHTED_EXCELS_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

# 修改 Excel 文件名，添加时间戳
output_filename = f"双色球结果和统计_{current_time}.xlsx"
output_path = EXCELS_DIR / output_filename

# 保存开奖结果
try:
    if output_path.exists():
        book = load_workbook(output_path)
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            df.to_excel(writer, sheet_name="开奖结果", index=False)
    else:
        df.to_excel(output_path, sheet_name="开奖结果", index=False)
    logging.info(f"开奖结果已保存到 {output_path}")
except (FileNotFoundError, InvalidFileException) as e:
    df.to_excel(output_path, sheet_name="开奖结果", index=False)
    logging.info(f"创建新文件并保存开奖结果: {output_path}")

# 统计和保存结果到一个 Excel 文件中的单个工作表
try:
    df = pd.read_excel(output_path, sheet_name="开奖结果")
except Exception as e:
    logging.error(f"读取文件时发生错误: {e}")
    sys.exit()

# 统计逻辑优化
red_counts = {i: {f"位置{j+1}": 0 for j in range(6)} for i in range(1, 34)}
blue_counts = {i: 0 for i in range(1, 17)}

# 统计星期二、四、日
weekdays = {"星期二": "二", "星期四": "四", "星期日": "日"}
red_counts_week = {day: {i: {f"位置{j+1}": 0 for j in range(6)} for i in range(1, 34)} for day in weekdays.values()}
blue_counts_week = {day: {i: 0 for i in range(1, 17)} for day in weekdays.values()}

# 统计每个数字的出现次数
for _, row in df.iterrows():
    red_numbers = row[3:9].astype(int).tolist()
    blue_number = int(row[9])
    week = row[2]

    for j, num in enumerate(red_numbers):
        if num in red_counts:
            red_counts[num][f"位置{j+1}"] += 1
        if week in weekdays.values() and num in red_counts_week[week]:
            red_counts_week[week][num][f"位置{j+1}"] += 1

    if blue_number in blue_counts:
        blue_counts[blue_number] += 1
    if week in weekdays.values() and blue_number in blue_counts_week[week]:
        blue_counts_week[week][blue_number] += 1

# 计算总期数
total_periods = len(df)
if total_periods == 0:
    logging.error("数据为空，没有统计结果。")
    sys.exit()

# 计算每个数字的概率并将其保存到 DataFrame
red_data = []
for num, counts in red_counts.items():
    row = {"号码": num}
    for pos, count in counts.items():
        row[f"{pos}次"] = count
        row[f"{pos}概率"] = count / total_periods
    red_data.append(row)

blue_data = []
for num, count in blue_counts.items():
    row = {"号码": num, "蓝球次数": count, "蓝球概率": count / total_periods}
    blue_data.append(row)

# 计算星期二、四、日的统计
red_data_week = {day: [] for day in weekdays.values()}
blue_data_week = {day: [] for day in weekdays.values()}

for day in weekdays.values():
    total_periods_week = len(df[df["星期"] == day])
    for num, counts in red_counts_week[day].items():
        row = {"号码": num}
        for pos, count in counts.items():
            row[f"{pos}次（星期{day}）"] = count
            row[f"{pos}概率（星期{day}）"] = count / total_periods_week if total_periods_week > 0 else 0
        red_data_week[day].append(row)

    for num, count in blue_counts_week[day].items():
        row = {"号码": num, f"蓝球次数（星期{day}）": count, f"蓝球概率（星期{day}）": count / total_periods_week if total_periods_week > 0 else 0}
        blue_data_week[day].append(row)

# 合并红球和蓝球数据
for red in red_data:
    num = red["号码"]
    blue_entry = next((b for b in blue_data if b["号码"] == num), None)
    if blue_entry:
        red["蓝球次数"] = blue_entry["蓝球次数"]
        red["蓝球概率"] = blue_entry["蓝球概率"]
    else:
        red["蓝球次数"] = 0
        red["蓝球概率"] = 0.0

for day in weekdays.values():
    for red in red_data_week[day]:
        num = red["号码"]
        blue_entry = next((b for b in blue_data_week[day] if b["号码"] == num), None)
        if blue_entry:
            red[f"蓝球次数（星期{day}）"] = blue_entry[f"蓝球次数（星期{day}）"]
            red[f"蓝球概率（星期{day}）"] = blue_entry[f"蓝球概率（星期{day}）"]
        else:
            red[f"蓝球次数（星期{day}）"] = 0
            red[f"蓝球概率（星期{day}）"] = 0.0

red_counts_df = pd.DataFrame(red_data)
red_counts_week_dfs = {day: pd.DataFrame(data) for day, data in red_data_week.items()}

# 修改统计结果的 sheet 名称，添加时间戳
statistics_sheet_name = f"统计结果_{current_time}"

try:
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        red_counts_df.to_excel(writer, sheet_name=statistics_sheet_name, index=False)
        for day, df_week in red_counts_week_dfs.items():
            sheet_name_week = f"星期{day}统计结果_{current_time}"
            df_week.to_excel(writer, sheet_name=sheet_name_week, index=False)
    logging.info(f"统计结果已保存到 {output_path}")
except Exception as e:
    logging.error(f"保存文件时发生错误: {e}")

# 高亮处理函数封装
# def highlight_max_min(sheet, col_start, col_end):
#     yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#     red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

#     for col in range(col_start, col_end + 1, 2):  # 假设概率列为偶数列
#         try:
#             column = [cell.value for cell in sheet.iter_cols(min_col=col, max_col=col,values_only=False)][0]
#             if not column:
#                 continue
#             max_value = max(column)
#             min_value = min(column)
#             for cell in sheet.iter_cols(min_col=col, max_col=col):
#                 for c in cell:
#                     if c.value == max_value:
#                         c.fill = red_fill
#                     elif c.value == min_value:
#                         c.fill = yellow_fill
#         except Exception as e:
#             logging.error(f"高亮处理列 {col} 时发生错误: {e}")
# def highlight_max_min(sheet, col_start, col_end):
#     from openpyxl.styles import PatternFill

#     yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#     red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

#     # 假如我们想跳过表头（第一行），我们可以指定 min_row=2
#     for col in range(col_start, col_end + 1, 2):
#         try:
#             # 1) 先获取该列所有的值（values_only=True），这样就可以做 max / min
#             col_values_gen = sheet.iter_cols(
#                 min_col=col, max_col=col,
#                 min_row=2,  # 如果第1行是表头
#                 values_only=True
#             )
            
#             # iter_cols(...) 返回一个生成器，其中仅有1个元素(因为max_col=col同min_col=col)
#             # 这个元素是一个元组，包含这一整列所有的值
#             col_values_tuple = next(col_values_gen)
            
#             # 如果整列都为空或都是表头，可能为空值
#             if not col_values_tuple or all(v is None for v in col_values_tuple):
#                 continue

#             # 排除掉为 None 或 字符串的情况，防止 max/min 出错
#             numeric_values = [v for v in col_values_tuple if isinstance(v, (int, float))]
#             if not numeric_values:
#                 continue

#             max_value = max(numeric_values)
#             min_value = min(numeric_values)

#             # 2) 再次获取该列的单元格对象（values_only=False），这样才能设置 fill
#             col_cells_gen = sheet.iter_cols(
#                 min_col=col, max_col=col,
#                 min_row=2,
#                 values_only=False
#             )
#             col_cells_tuple = next(col_cells_gen)  # 一列的所有 Cell 对象

#             for cell in col_cells_tuple:
#                 cell_value = cell.value
#                 if cell_value == max_value:
#                     cell.fill = red_fill
#                 elif cell_value == min_value:
#                     cell.fill = yellow_fill

#         except Exception as e:
#             print(f"高亮处理列 {col} 时发生错误: {e}")


# 高亮处理
# try:
#     book = load_workbook(output_path)
#     sheet = book[statistics_sheet_name]
#     # 假设概率列从第4列开始
#     highlight_max_min(sheet, 4, sheet.max_column)
#     # 修改高亮处理后的文件名，添加时间戳
#     highlighted_output_filename = f"双色球结果和统计_highlighted_{current_time}.xlsx"
#     highlighted_output_path = HIGHLIGHTED_EXCELS_DIR / highlighted_output_filename
#     book.save(highlighted_output_path)
#     logging.info(f"高亮处理已完成并保存到 {highlighted_output_path}")
# except Exception as e:
#     logging.error(f"高亮处理时发生错误: {e}")

# 定义高亮处理函数
def highlight_max_min(sheet, col_start, col_end):
    from openpyxl.styles import PatternFill

    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col in range(col_start, col_end + 1, 2):
        try:
            # 获取该列的所有值，跳过表头
            col_values_gen = sheet.iter_cols(
                min_col=col, max_col=col,
                min_row=2,
                values_only=True
            )
            col_values_tuple = next(col_values_gen)

            # 过滤非数字值
            numeric_values = [v for v in col_values_tuple if isinstance(v, (int, float))]
            if not numeric_values:
                continue

            max_value = max(numeric_values)
            min_value = min(numeric_values)

            # 获取单元格对象以应用填充
            col_cells_gen = sheet.iter_cols(
                min_col=col, max_col=col,
                min_row=2,
                values_only=False
            )
            col_cells_tuple = next(col_cells_gen)

            for cell in col_cells_tuple:
                cell_value = cell.value
                if cell_value == max_value:
                    cell.fill = RED_FILL
                elif cell_value == min_value:
                    cell.fill = YELLOW_FILL

        except Exception as e:
            logging.error(f"高亮处理列 {col} 时发生错误: {e}")

# 定义格式化 "开奖结果" 工作表的函数
def format_results_sheet(sheet):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    RED_FONT = Font(color="FF0000")
    BLUE_FONT = Font(color="0000FF")
    LIGHT_RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    
    # 假设红球号码列为 "红球号码1" 到 "红球号码6"（列 D 到 I）
    red_ball_columns = ["D", "E", "F", "G", "H", "I"]
    # 假设蓝球号码列为 "蓝球号码"（列 J）
    blue_ball_column = "J"

    # 设置红球和蓝球列的字体颜色
    for col in red_ball_columns:
        for cell in sheet[col][1:]:  # 跳过表头
            cell.font = RED_FONT
    for cell in sheet[blue_ball_column][1:]:
        cell.font = BLUE_FONT

    # 应用交替行背景颜色
    for row in range(2, sheet.max_row + 1):
        if row % 2 == 0:
            # 偶数行应用浅红和浅蓝背景
            for col in red_ball_columns:
                sheet[f"{col}{row}"].fill = LIGHT_RED_FILL
            sheet[f"{blue_ball_column}{row}"].fill = LIGHT_BLUE_FILL
        else:
            # 奇数行保持默认背景
            pass  # 如果需要，可以设置为其他颜色

# ... [数据获取和处理部分保持不变] ...

# 保存统计结果到 Excel
try:
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        red_counts_df.to_excel(writer, sheet_name=statistics_sheet_name, index=False)
        for day, df_week in red_counts_week_dfs.items():
            sheet_name_week = f"星期{day}统计结果_{current_time}"
            df_week.to_excel(writer, sheet_name=sheet_name_week, index=False)
    logging.info(f"统计结果已保存到 {output_path}")
except (FileNotFoundError, InvalidFileException) as e:
    df.to_excel(output_path, sheet_name="开奖结果", index=False)
    logging.info(f"创建新文件并保存开奖结果: {output_path}")

# 高亮处理和格式化 "开奖结果" 工作表
try:
    book = load_workbook(output_path)
    
    # 定义所有需要高亮处理的工作表名称
    sheets_to_highlight = [statistics_sheet_name] + [
        f"星期{day}统计结果_{current_time}" for day in weekdays.values()
    ]
    
    for sheet_name in sheets_to_highlight:
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            # 假设概率列从第4列开始
            highlight_max_min(sheet, 2, sheet.max_column)
            logging.info(f"高亮处理已完成: {sheet_name}")
        else:
            logging.warning(f"工作表 '{sheet_name}' 未找到，跳过高亮处理。")
    
    # 格式化 "开奖结果" 工作表
    if "开奖结果" in book.sheetnames:
        results_sheet = book["开奖结果"]
        format_results_sheet(results_sheet)
        logging.info("格式化 '开奖结果' 工作表完成。")
    else:
        logging.error("工作表 '开奖结果' 未找到，无法进行格式化。")
    
    # 修改高亮和格式化处理后的文件名，添加时间戳
    final_output_filename = f"双色球结果和统计_final_{current_time}.xlsx"
    final_output_path = HIGHLIGHTED_EXCELS_DIR / final_output_filename
    book.save(final_output_path)
    logging.info(f"所有高亮和格式化处理已完成并保存到 {final_output_path}")
except Exception as e:
    logging.error(f"高亮和格式化处理时发生错误: {e}")
#-------------------------------------------------------------------------------------------------
# 可视化部分封装
def save_heatmap(data, title, filepath, cmap="YlGnBu"):
    plt.figure(figsize=(12, 8))
    sns.heatmap(data, annot=True, fmt="d", cmap=cmap)
    plt.title(title)
    plt.savefig(filepath)
    plt.close()

def save_distribution(data, title, xlabel, ylabel, filepath):
    plt.figure(figsize=(12, 8))
    sns.barplot(x=data.index, y=data.values)
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.savefig(filepath)
    plt.close()

# 创建热力图和分布图的函数
def create_visualizations():
    try:
        # 创建红球号码出现频率的热力图
        red_heatmap_data = pd.DataFrame(red_counts).T.astype(int)
        red_heatmap_filename = f"红球号码出现频率热力图_{current_time}.png"
        red_heatmap_path = RED_PLOTS_DIR / red_heatmap_filename
        save_heatmap(red_heatmap_data, '红球号码热力图', red_heatmap_path)

        # 创建蓝球号码出现频率的热力图
        blue_heatmap_data = pd.DataFrame.from_dict(blue_counts, orient='index', columns=['次数']).astype(int)
        blue_heatmap_filename = f"蓝球号码出现频率热力图_{current_time}.png"
        blue_heatmap_path = BLUE_PLOTS_DIR / blue_heatmap_filename
        save_heatmap(blue_heatmap_data.T, '蓝球号码热力图', blue_heatmap_path, cmap="YlOrBr")

        # 创建红球号码出现频率的分布图
        red_numbers_total_counts = {i: sum(pos_counts.values()) for i, pos_counts in red_counts.items()}
        red_numbers_series = pd.Series(red_numbers_total_counts).sort_index().astype(int)
        red_distribution_filename = f"红球号码分布_{current_time}.png"
        red_distribution_path = RED_PLOTS_DIR / red_distribution_filename
        save_distribution(red_numbers_series, '红球号码分布', '号码', '频率', red_distribution_path)

        # 创建蓝球号码出现频率的分布图
        blue_numbers_series = pd.Series(blue_counts).sort_index().astype(int)
        blue_distribution_filename = f"蓝球号码分布_{current_time}.png"
        blue_distribution_path = BLUE_PLOTS_DIR / blue_distribution_filename
        save_distribution(blue_numbers_series, '蓝球号码分布', '号码', '频率', blue_distribution_path)

        logging.info("所有可视化图表已创建并保存。")
    except Exception as e:
        logging.error(f"可视化创建时发生错误: {e}")

# 使用线程池并行生成可视化图表
with ThreadPoolExecutor(max_workers=4) as executor:
    executor.submit(create_visualizations)

# 推荐号码生成部分保持不变
# （根据之前的代码，你可能需要将生成的推荐号码输出路径也调整到相应目录中）

# 调用预测脚本
def run_prediction_script(excel_file, output_dir):
    try:
        prediction_script = PROJECT_ROOT / "time_series_prediction_with_all.py"
        if not prediction_script.exists():
            logging.error(f"预测脚本 '{prediction_script}' 未找到。")
            return

        # 传递 output_dir 作为参数
        subprocess.run([sys.executable, str(prediction_script), str(excel_file), str(output_dir)], check=True)
        logging.info("预测脚本已成功运行。")
    except subprocess.CalledProcessError as e:
        logging.error(f"运行预测脚本时发生错误: {e}")
    except Exception as e:
        logging.error(f"发生错误: {e}")

# 在数据处理完毕后调用预测脚本，并传递新文件夹路径
run_prediction_script(output_path, new_folder_path)

# 结束计时并计算总耗时
end_time = time.time()
duration = end_time - start_time
logging.info(f"程序总耗时: {duration:.2f} 秒")
